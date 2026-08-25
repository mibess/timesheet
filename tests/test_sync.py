from __future__ import annotations

import shutil
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from timesheet_ccee.database import TimesheetDatabase
from timesheet_ccee.models import TimeEntry, WorkbookMetadata
from timesheet_ccee.sync import TimesheetSync
from timesheet_ccee.workbook import DATA_SHEET, TimesheetError, TimesheetWorkbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent


class TimesheetSyncTests(unittest.TestCase):
    def test_reset_replaces_workbook_and_clears_database(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            workbook_path = root / "timesheet.xlsx"
            template_path = PROJECT_ROOT / "Modelo_Timesheet_CCEE.template.xlsx"
            shutil.copy2(template_path, workbook_path)
            database = TimesheetDatabase(root / "timesheet.sqlite3")
            synchronizer = TimesheetSync(database)
            workbook = TimesheetWorkbook(workbook_path)
            synchronizer.activate(workbook)
            database.save_day(
                workbook_path,
                date(2099, 1, 2),
                [TimeEntry("01:00", "Sustentação", "CSTM", "123")],
            )

            outcome = synchronizer.reset(workbook, template_path)

            self.assertTrue(outcome.synchronized)
            self.assertEqual(outcome.record_count, 0)
            self.assertTrue(Path(outcome.backup_path).is_file())
            self.assertEqual(database.load_all_entries(workbook_path), [])
            self.assertEqual(workbook.load_dataset()[1], [])
            self.assertFalse(database.needs_sync(workbook_path))

    def test_exports_empty_number_as_null_cell(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            workbook_path = Path(temporary_directory) / "timesheet.xlsx"
            shutil.copy2(
                PROJECT_ROOT / "Modelo_Timesheet_CCEE.template.xlsx",
                workbook_path,
            )
            workbook = TimesheetWorkbook(workbook_path)

            workbook.save_all(
                [
                    (
                        date(2099, 1, 2),
                        TimeEntry("01:00", "Sustentação", "CSTM", "   "),
                    )
                ]
            )

            saved_workbook = load_workbook(workbook_path, data_only=True)
            try:
                self.assertIsNone(saved_workbook[DATA_SHEET].cell(2, 5).value)
            finally:
                saved_workbook.close()

    def test_first_use_imports_and_later_sync_exports_database(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            workbook_path = root / "timesheet.xlsx"
            shutil.copy2(
                PROJECT_ROOT / "Modelo_Timesheet_CCEE.template.xlsx",
                workbook_path,
            )
            database = TimesheetDatabase(root / "timesheet.sqlite3")
            synchronizer = TimesheetSync(database)
            workbook = TimesheetWorkbook(workbook_path)

            initial_entries = workbook.load_dataset()[1]
            imported = synchronizer.activate(workbook)
            self.assertTrue(imported.imported)
            self.assertEqual(imported.record_count, len(initial_entries))
            self.assertFalse(database.needs_sync(workbook_path))

            selected_date = date(2099, 1, 2)
            entry = TimeEntry(
                "03:45", "Sustentação", "CSTM", "999", "Teste de sincronização"
            )
            database.save_day(workbook_path, selected_date, [entry])
            synchronized = synchronizer.synchronize(workbook)

            self.assertTrue(synchronized.synchronized)
            self.assertEqual(synchronized.record_count, len(initial_entries) + 1)
            self.assertTrue(Path(synchronized.backup_path).is_file())
            self.assertFalse(database.needs_sync(workbook_path))
            self.assertEqual(workbook.load_day(selected_date), [entry])

            unchanged = synchronizer.synchronize(workbook)
            self.assertFalse(unchanged.synchronized)
            self.assertEqual(unchanged.backup_path, "")

    def test_import_data_replaces_database_and_synchronizes_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            workbook_path = root / "timesheet.xlsx"
            source_path = root / "importacao.xlsx"
            template_path = PROJECT_ROOT / "Modelo_Timesheet_CCEE.template.xlsx"
            shutil.copy2(template_path, workbook_path)
            shutil.copy2(template_path, source_path)
            database = TimesheetDatabase(root / "timesheet.sqlite3")
            synchronizer = TimesheetSync(database)
            workbook = TimesheetWorkbook(workbook_path)
            source = TimesheetWorkbook(source_path)
            synchronizer.activate(workbook)
            database.save_day(
                workbook_path,
                date(2026, 8, 24),
                [TimeEntry("01:00", "ADM", "Reuniões")],
            )
            imported_entry = TimeEntry(
                "02:30", "Importada", "Novo ticket", "321", phase="Nova fase"
            )
            imported_metadata = WorkbookMetadata(
                activity_types=["Importada"],
                ticket_types=["Novo ticket"],
                phases=["Nova fase"],
                recent_numbers=["321"],
            )
            source.save_all(
                [(date(2026, 8, 25), imported_entry)],
                metadata=imported_metadata,
            )

            outcome = synchronizer.import_data(workbook, source)

            self.assertTrue(outcome.imported)
            self.assertTrue(outcome.synchronized)
            self.assertEqual(outcome.record_count, 1)
            self.assertTrue(Path(outcome.backup_path).is_file())
            self.assertEqual(
                database.load_all_entries(workbook_path),
                [(date(2026, 8, 25), imported_entry)],
            )
            self.assertEqual(database.load_metadata(workbook_path), imported_metadata)
            self.assertEqual(workbook.load_dataset()[1], source.load_dataset()[1])
            self.assertEqual(workbook.load_metadata(), imported_metadata)
            self.assertFalse(database.needs_sync(workbook_path))

    def test_import_data_rejects_invalid_source_without_changing_database(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            workbook_path = root / "timesheet.xlsx"
            invalid_source_path = root / "invalida.xlsx"
            shutil.copy2(
                PROJECT_ROOT / "Modelo_Timesheet_CCEE.template.xlsx",
                workbook_path,
            )
            invalid_source_path.write_bytes(b"nao e uma planilha")
            database = TimesheetDatabase(root / "timesheet.sqlite3")
            synchronizer = TimesheetSync(database)
            workbook = TimesheetWorkbook(workbook_path)
            synchronizer.activate(workbook)
            original_entry = TimeEntry("01:00", "ADM", "Reuniões")
            original_date = date(2026, 8, 24)
            database.save_day(workbook_path, original_date, [original_entry])

            with self.assertRaises(TimesheetError):
                synchronizer.import_data(
                    workbook, TimesheetWorkbook(invalid_source_path)
                )

            self.assertEqual(
                database.load_all_entries(workbook_path),
                [(original_date, original_entry)],
            )


if __name__ == "__main__":
    unittest.main()
