from __future__ import annotations

import os
import re
import shutil
import tempfile
import warnings
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from .models import SaveResult, TimeEntry, WorkbookMetadata
from .time_utils import (
    excel_duration_to_minutes,
    excel_value_to_date,
    format_duration,
    parse_duration,
)


DATA_SHEET = "Apontamento Terceiros"
LIST_SHEET = "Lista de dados"
FIRST_DATA_ROW = 2
DEFAULT_LAST_DATA_ROW = 312
EXPECTED_HEADERS = (
    "Data Trabalhada",
    "Horas Apontadas",
    "Tipo de Atividade",
    "Ticket",
    "Número",
    "Observação",
)

_EXTENSION_PATTERN = re.compile(rb"<extLst(?:\s[^>]*)?>.*?</extLst>", re.DOTALL)
_XR_NAMESPACE = b'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision"'


class TimesheetError(RuntimeError):
    """Erro seguro para exibição na interface."""


@dataclass
class _Record:
    worked_date: date
    hours_fraction: float
    activity_type: str
    ticket: str
    number: str | int
    observation: str
    phase: str
    order: int


def _unique_text(values: Iterable[object]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = "" if value is None else str(value).strip()
        if text and text not in seen:
            result.append(text)
            seen.add(text)
    return result


def _number_for_excel(value: str) -> str | int:
    text = value.strip()
    return int(text) if text.isdigit() else text


def _load_excel(path: Path, **options):
    # O modelo usa validações x14, que o openpyxl lê, mas não reescreve.
    # A extensão original é restaurada após o save em _restore_sheet_extensions.
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
        )
        return load_workbook(path, **options)


def _worksheet_extensions(path: Path) -> dict[str, bytes]:
    extensions: dict[str, bytes] = {}
    with zipfile.ZipFile(path, "r") as archive:
        for name in archive.namelist():
            if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
                continue
            payload = archive.read(name)
            match = _EXTENSION_PATTERN.search(payload)
            if not match:
                continue
            fragment = match.group(0)
            if b"xr:" in fragment and b"xmlns:xr=" not in fragment:
                fragment = fragment.replace(
                    b"<extLst>", b"<extLst " + _XR_NAMESPACE + b">", 1
                )
            extensions[name] = fragment
    return extensions


def _restore_sheet_extensions(path: Path, extensions: dict[str, bytes]) -> None:
    if not extensions:
        return
    descriptor, patched_name = tempfile.mkstemp(
        prefix=f".{path.stem}_extensions_", suffix=path.suffix, dir=path.parent
    )
    os.close(descriptor)
    patched_path = Path(patched_name)
    try:
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
            patched_path, "w"
        ) as destination:
            for info in source.infolist():
                payload = source.read(info.filename)
                fragment = extensions.get(info.filename)
                if fragment:
                    payload = _EXTENSION_PATTERN.sub(b"", payload)
                    payload = payload.replace(
                        b"</worksheet>", fragment + b"</worksheet>", 1
                    )
                destination.writestr(info, payload)
        os.replace(patched_path, path)
    finally:
        patched_path.unlink(missing_ok=True)


class TimesheetWorkbook:
    def __init__(self, path: str | Path):
        self.path = Path(path).expanduser().resolve()

    @property
    def keep_vba(self) -> bool:
        return self.path.suffix.lower() == ".xlsm"

    def validate(self) -> None:
        if not self.path.is_file():
            raise TimesheetError("A planilha selecionada não foi encontrada.")
        if self.path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise TimesheetError("Selecione uma planilha .xlsx ou .xlsm.")

        try:
            workbook = _load_excel(
                self.path,
                read_only=True,
                data_only=False,
                keep_links=True,
                keep_vba=self.keep_vba,
            )
        except Exception as exc:
            raise TimesheetError(
                "Não foi possível abrir a planilha. Verifique se o arquivo é válido."
            ) from exc

        try:
            missing = {DATA_SHEET, LIST_SHEET}.difference(workbook.sheetnames)
            if missing:
                raise TimesheetError(
                    "A planilha não segue o modelo CCEE. Aba ausente: "
                    + ", ".join(sorted(missing))
                )
            sheet = workbook[DATA_SHEET]
            headers = tuple(sheet.cell(1, column).value for column in range(1, 7))
            if headers != EXPECTED_HEADERS:
                raise TimesheetError(
                    "Os cabeçalhos da aba de apontamentos não correspondem ao modelo CCEE."
                )
        finally:
            workbook.close()

    def load_metadata(self) -> WorkbookMetadata:
        self.validate()
        workbook = _load_excel(
            self.path,
            read_only=True,
            data_only=True,
            keep_links=True,
            keep_vba=self.keep_vba,
        )
        try:
            return self._metadata_from_workbook(workbook)
        finally:
            workbook.close()

    def load_day(self, selected_date: date) -> list[TimeEntry]:
        self.validate()
        workbook = _load_excel(
            self.path,
            read_only=True,
            data_only=True,
            keep_links=True,
            keep_vba=self.keep_vba,
        )
        try:
            return self._entries_from_workbook(workbook, selected_date)
        finally:
            workbook.close()

    def load_snapshot(
        self, selected_date: date
    ) -> tuple[WorkbookMetadata, list[TimeEntry]]:
        """Carrega opções e atividades usando uma única abertura de leitura."""
        self.validate()
        workbook = _load_excel(
            self.path,
            read_only=True,
            data_only=True,
            keep_links=True,
            keep_vba=self.keep_vba,
        )
        try:
            return (
                self._metadata_from_workbook(workbook),
                self._entries_from_workbook(workbook, selected_date),
            )
        finally:
            workbook.close()

    def load_dataset(self) -> tuple[WorkbookMetadata, list[tuple[date, TimeEntry]]]:
        """Lê todos os dados necessários para a importação inicial no SQLite."""
        self.validate()
        workbook = _load_excel(
            self.path,
            read_only=True,
            data_only=True,
            keep_links=True,
            keep_vba=self.keep_vba,
        )
        try:
            return (
                self._metadata_from_workbook(workbook),
                self._dated_entries_from_workbook(workbook),
            )
        finally:
            workbook.close()

    @staticmethod
    def _metadata_from_workbook(workbook) -> WorkbookMetadata:
        list_sheet = workbook[LIST_SHEET]
        data_sheet = workbook[DATA_SHEET]
        activity_types = _unique_text(
            list_sheet.cell(row, 3).value for row in range(4, 21)
        )
        ticket_types = _unique_text(
            list_sheet.cell(row, 4).value for row in range(4, 21)
        )
        phases = _unique_text(
            list_sheet.cell(row, 5).value for row in range(4, 21)
        )

        recent_numbers: list[str] = []
        seen: set[str] = set()
        for row in range(DEFAULT_LAST_DATA_ROW, FIRST_DATA_ROW - 1, -1):
            value = data_sheet.cell(row, 5).value
            if value in (None, ""):
                continue
            if isinstance(value, float) and value.is_integer():
                text = str(int(value))
            else:
                text = str(value).strip()
            if text and text not in seen:
                recent_numbers.append(text)
                seen.add(text)
            if len(recent_numbers) == 15:
                break

        return WorkbookMetadata(
            activity_types=activity_types,
            ticket_types=ticket_types,
            phases=phases,
            recent_numbers=recent_numbers,
        )

    @staticmethod
    def _entries_from_workbook(workbook, selected_date: date) -> list[TimeEntry]:
        return [
            entry
            for worked_date, entry in TimesheetWorkbook._dated_entries_from_workbook(
                workbook
            )
            if worked_date == selected_date
        ]

    @staticmethod
    def _dated_entries_from_workbook(workbook) -> list[tuple[date, TimeEntry]]:
        sheet = workbook[DATA_SHEET]
        entries: list[tuple[date, TimeEntry]] = []
        for row in range(FIRST_DATA_ROW, DEFAULT_LAST_DATA_ROW + 1):
            worked_date = excel_value_to_date(sheet.cell(row, 1).value)
            if worked_date is None:
                continue
            number_value = sheet.cell(row, 5).value
            if isinstance(number_value, float) and number_value.is_integer():
                number = str(int(number_value))
            else:
                number = "" if number_value is None else str(number_value)
            entries.append(
                (
                    worked_date,
                    TimeEntry(
                        hours=format_duration(
                            excel_duration_to_minutes(sheet.cell(row, 2).value)
                        ),
                        activity_type=str(sheet.cell(row, 3).value or ""),
                        ticket=str(sheet.cell(row, 4).value or ""),
                        number=number,
                        observation=str(sheet.cell(row, 6).value or ""),
                        phase=str(sheet.cell(row, 7).value or ""),
                    ),
                )
            )
        return entries

    @staticmethod
    def _last_data_row(sheet) -> int:
        for table in sheet.tables.values():
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if min_col == 1 and min_row == 1 and max_col >= 7:
                return max_row
        return DEFAULT_LAST_DATA_ROW

    @staticmethod
    def _check_file_access(path: Path) -> None:
        try:
            with path.open("r+b"):
                pass
        except (PermissionError, OSError) as exc:
            raise TimesheetError(
                "A planilha está aberta ou bloqueada. Feche o arquivo no Excel "
                "e tente novamente."
            ) from exc

    def save_all(self, entries: Iterable[tuple[date, TimeEntry]]) -> SaveResult:
        """Substitui os apontamentos da planilha pelo estado integral do banco."""
        self.validate()
        self._check_file_access(self.path)
        normalized: list[_Record] = []
        date_orders: dict[date, int] = {}
        for row_number, (worked_date, entry) in enumerate(entries, start=1):
            order = date_orders.get(worked_date, 0) + 1
            date_orders[worked_date] = order
            try:
                minutes = parse_duration(entry.hours, allow_zero=True)
            except ValueError as exc:
                raise TimesheetError(f"Registro {row_number}: {exc}") from exc
            activity_type = entry.activity_type.strip()
            ticket = entry.ticket.strip()
            if not activity_type:
                raise TimesheetError(
                    f"Informe o Tipo de Atividade no registro {row_number}."
                )
            if not ticket:
                raise TimesheetError(f"Informe o Ticket no registro {row_number}.")
            normalized.append(
                _Record(
                    worked_date=worked_date,
                    hours_fraction=minutes / 1440,
                    activity_type=activity_type,
                    ticket=ticket,
                    number=_number_for_excel(entry.number),
                    observation=entry.observation.strip(),
                    phase=entry.phase.strip(),
                    order=order,
                )
            )

        records = sorted(
            normalized, key=lambda record: (record.worked_date, record.order)
        )
        backup_dir = self.path.parent / "backups-timesheet"
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        backup_path = backup_dir / (
            f"{self.path.stem}_backup_{timestamp}{self.path.suffix}"
        )
        shutil.copy2(self.path, backup_path)
        original_extensions = _worksheet_extensions(self.path)

        workbook = _load_excel(
            self.path,
            data_only=False,
            keep_links=True,
            keep_vba=self.keep_vba,
        )
        temp_path: Path | None = None
        try:
            sheet = workbook[DATA_SHEET]
            last_row = self._last_data_row(sheet)
            capacity = last_row - FIRST_DATA_ROW + 1
            if len(records) > capacity:
                raise TimesheetError(
                    f"A planilha atingiu o limite de {capacity} registros."
                )

            for row in range(FIRST_DATA_ROW, last_row + 1):
                for column in range(1, 8):
                    sheet.cell(row, column).value = None

            for row, record in enumerate(records, start=FIRST_DATA_ROW):
                values = (
                    record.worked_date,
                    record.hours_fraction,
                    record.activity_type,
                    record.ticket,
                    record.number,
                    record.observation,
                    record.phase,
                )
                for column, value in enumerate(values, start=1):
                    sheet.cell(row, column).value = value
                sheet.cell(row, 1).number_format = "dd/mm/yyyy"
                sheet.cell(row, 2).number_format = "[h]:mm"

            descriptor, temp_name = tempfile.mkstemp(
                prefix=f".{self.path.stem}_",
                suffix=self.path.suffix,
                dir=self.path.parent,
            )
            os.close(descriptor)
            temp_path = Path(temp_name)
            workbook.save(temp_path)
            _restore_sheet_extensions(temp_path, original_extensions)
            shutil.copymode(self.path, temp_path)
            os.replace(temp_path, self.path)
            temp_path = None
            return SaveResult(str(backup_path), len(records))
        except TimesheetError:
            raise
        except PermissionError as exc:
            raise TimesheetError(
                "Não foi possível substituir a planilha. Feche o arquivo no Excel "
                "e tente novamente."
            ) from exc
        except Exception as exc:
            raise TimesheetError(
                "Não foi possível sincronizar a planilha. O backup original foi preservado."
            ) from exc
        finally:
            workbook.close()
            if temp_path is not None:
                try:
                    temp_path.unlink(missing_ok=True)
                except OSError:
                    pass

    def save_day(self, selected_date: date, entries: list[TimeEntry]) -> SaveResult:
        self.validate()
        self._check_file_access(self.path)
        normalized: list[_Record] = []
        for order, entry in enumerate(entries, start=1):
            try:
                minutes = parse_duration(entry.hours, allow_zero=True)
            except ValueError as exc:
                raise TimesheetError(f"Linha {order}: {exc}") from exc
            if minutes == 0:
                raise TimesheetError(
                    f"Linha {order}: as horas estão em 00:00. Edite essa atividade "
                    "ou remova a linha antes de salvar."
                )
            activity_type = entry.activity_type.strip()
            ticket = entry.ticket.strip()
            if not activity_type:
                raise TimesheetError(
                    f"Informe o Tipo de Atividade na linha {order}."
                )
            if not ticket:
                raise TimesheetError(f"Informe o Ticket na linha {order}.")
            normalized.append(
                _Record(
                    worked_date=selected_date,
                    hours_fraction=minutes / 1440,
                    activity_type=activity_type,
                    ticket=ticket,
                    number=_number_for_excel(entry.number),
                    observation=entry.observation.strip(),
                    phase=entry.phase.strip(),
                    order=order,
                )
            )

        backup_dir = self.path.parent / "backups-timesheet"
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        backup_path = backup_dir / (
            f"{self.path.stem}_backup_{timestamp}{self.path.suffix}"
        )
        shutil.copy2(self.path, backup_path)
        original_extensions = _worksheet_extensions(self.path)

        workbook = _load_excel(
            self.path,
            data_only=False,
            keep_links=True,
            keep_vba=self.keep_vba,
        )
        temp_path: Path | None = None
        try:
            sheet = workbook[DATA_SHEET]
            last_row = self._last_data_row(sheet)
            existing: list[_Record] = []
            order = 0
            for row in range(FIRST_DATA_ROW, last_row + 1):
                worked_date = excel_value_to_date(sheet.cell(row, 1).value)
                if worked_date is None or worked_date == selected_date:
                    continue
                order += 1
                number_value = sheet.cell(row, 5).value
                if isinstance(number_value, float) and number_value.is_integer():
                    number_value = int(number_value)
                existing.append(
                    _Record(
                        worked_date=worked_date,
                        hours_fraction=excel_duration_to_minutes(
                            sheet.cell(row, 2).value
                        )
                        / 1440,
                        activity_type=str(sheet.cell(row, 3).value or ""),
                        ticket=str(sheet.cell(row, 4).value or ""),
                        number="" if number_value is None else number_value,
                        observation=str(sheet.cell(row, 6).value or ""),
                        phase=str(sheet.cell(row, 7).value or ""),
                        order=order,
                    )
                )

            records = sorted(
                [*existing, *normalized],
                key=lambda record: (record.worked_date, record.order),
            )
            capacity = last_row - FIRST_DATA_ROW + 1
            if len(records) > capacity:
                raise TimesheetError(
                    f"A planilha atingiu o limite de {capacity} registros."
                )

            for row in range(FIRST_DATA_ROW, last_row + 1):
                for column in range(1, 8):
                    sheet.cell(row, column).value = None

            for row, record in enumerate(records, start=FIRST_DATA_ROW):
                values = (
                    record.worked_date,
                    record.hours_fraction,
                    record.activity_type,
                    record.ticket,
                    record.number,
                    record.observation,
                    record.phase,
                )
                for column, value in enumerate(values, start=1):
                    sheet.cell(row, column).value = value
                sheet.cell(row, 1).number_format = "dd/mm/yyyy"
                sheet.cell(row, 2).number_format = "[h]:mm"

            descriptor, temp_name = tempfile.mkstemp(
                prefix=f".{self.path.stem}_",
                suffix=self.path.suffix,
                dir=self.path.parent,
            )
            os.close(descriptor)
            temp_path = Path(temp_name)
            workbook.save(temp_path)
            _restore_sheet_extensions(temp_path, original_extensions)
            shutil.copymode(self.path, temp_path)
            os.replace(temp_path, self.path)
            temp_path = None
            return SaveResult(str(backup_path), len(records))
        except TimesheetError:
            raise
        except PermissionError as exc:
            raise TimesheetError(
                "Não foi possível substituir a planilha. Feche o arquivo no Excel "
                "e tente novamente."
            ) from exc
        except Exception as exc:
            raise TimesheetError(
                "Não foi possível salvar a planilha. O backup original foi preservado."
            ) from exc
        finally:
            workbook.close()
            if temp_path is not None:
                try:
                    temp_path.unlink(missing_ok=True)
                except OSError:
                    pass
